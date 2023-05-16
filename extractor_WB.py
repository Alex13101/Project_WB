
import requests
import time
import datetime
from datetime import date
import openpyxl

def get_products(url):
    products = requests.get(url).json()['data']['products']
    return products

search = input('Что будем искать? ')




start = time.time()
def search_right_product(search):
    products = []
    urls = [
        f'https://search.wb.ru/exactmatch/ru/common/v4/search?appType=1&curr=rub&dest=-1281648&page={i}&query={search}&regions=80,64,38,4,115,83,33,68,70,69,30,86,75,40,1,66,48,110,31,22,71,114&resultset=catalog&sort=popular&spp=22&suppressSpellcheck=false'
        for i in range(1, 61)]
    for url in urls:
        products += get_products(url)

    return products


# print(len(all_products))
# print(len(all_products[1]))


long_lst = len(search_right_product(search))
all_products = search_right_product(search)

excel_file = openpyxl.load_workbook('Base.xlsx')  # Открытие exel-файла
sheet_sheet = excel_file['Лист1']
current_date = date.today()
current_date_time = datetime.datetime.now()
current_time = current_date_time.time()

for i in range(0, long_lst):
    title = all_products[i].setdefault("name")
    price = all_products[i].setdefault('salePriceU')
    id_product = all_products[i].setdefault('id')
    brand_product = all_products[i].setdefault('brand')
    print(title, 'стоит ', price/100, ' рублей')

    sheet_sheet.cell(row=i + 3, column=2).value = id_product
    sheet_sheet.cell(row=i + 3, column=3).value = title
    sheet_sheet.cell(row=i + 3, column=4).value = brand_product
    sheet_sheet.cell(row=i + 3, column=5).value = price/100
    sheet_sheet.cell(row=i + 3, column=6).value = current_date
    sheet_sheet.cell(row=i + 3, column=7).value = current_time
    #print( 'Готово!')
excel_file.save('Base.xlsx')

excel_file.close()

print(time.time() - start)

# for k, v in all_products[15].items():
#     print(k, v)
